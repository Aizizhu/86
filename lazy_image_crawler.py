"""XC8866 crawler that combines list-page crawling with rendered lazy-image extraction.

This module keeps the original workflow requested by the user: paginated list
URLs, progress.json checkpointing, image download, Excel rows with embedded
images, and retry handling. Topic parsing is upgraded to render pages in
Playwright before extracting images, which is necessary for Vue/Element Plus
lazy-loaded markup such as ``.topic-detail-image img.el-image__preview``.
"""

from __future__ import annotations

import argparse
import csv
import json
import random
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable, Sequence
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage

BASE_DOMAIN = "https://xc8866.com"
BASE_URL = f"{BASE_DOMAIN}/topic/{{topic_id:06d}}"
PROGRESS_FILE = "progress.json"
IMAGE_DIR = "images"
HEADERS = {"User-Agent": "Mozilla/5.0", "Referer": BASE_DOMAIN}
DEFAULT_WAIT_SELECTOR = "img.el-image__preview, .topic-detail-image img, .el-image img"
IMAGE_SELECTORS = (
    ".topic-detail-image img",
    "img.el-image__preview",
    ".el-image img",
    "img[src]",
)
PLACEHOLDER_IMAGE_MARKERS = ("zwzp.jpg", "default.jpg", "nopic.jpg")
ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")
FILENAME_ILLEGAL_RE = re.compile(r'[\\/:*?"<>|]')

session = requests.Session()
session.headers.update(HEADERS)


@dataclass
class TopicResult:
    topic_id: str
    url: str
    ok: bool
    title: str = ""
    price: str = ""
    address: str = ""
    qq: str = ""
    wechat: str = ""
    phone: str = ""
    content: str = ""
    image_count: int = 0
    image_urls: str = ""
    error: str = ""

    @property
    def image_list(self) -> list[str]:
        return [item for item in self.image_urls.splitlines() if item]


def clean_text(text: Any) -> str:
    if not text:
        return ""
    return ILLEGAL_CHARACTERS_RE.sub("", str(text))


def clean_filename(name: str) -> str:
    return FILENAME_ILLEGAL_RE.sub("_", clean_text(name))[:120] or "untitled"


def build_topic_url(topic_id: int | str) -> str:
    return BASE_URL.format(topic_id=int(topic_id))


def build_page_url(base_url: str, page_num: int) -> str:
    if "page=" in base_url:
        return re.sub(r"page=\d+", f"page={page_num}", base_url)
    if "?" in base_url:
        return f"{base_url}&page={page_num}"
    return f"{base_url}?page={page_num}"


def parse_topic_id(url: str) -> str:
    match = re.search(r"/topic/(\d+)", url)
    return match.group(1) if match else ""


def load_progress(progress_file: Path) -> set[str]:
    if progress_file.exists():
        return set(json.loads(progress_file.read_text(encoding="utf-8")))
    return set()


def save_progress(done_pages: set[str], progress_file: Path) -> None:
    progress_file.write_text(json.dumps(sorted(done_pages), ensure_ascii=False, indent=2), encoding="utf-8")


def request_url(url: str, timeout: int = 15) -> requests.Response | None:
    try:
        response = session.get(url, timeout=timeout)
        response.raise_for_status()
        return response
    except requests.RequestException:
        return None


def is_valid_topic_image(src: str, classes: Sequence[str] | str | None = None) -> bool:
    if not src or not src.startswith("http"):
        return False
    if "/avatars/" in src:
        return False
    if any(marker in src for marker in PLACEHOLDER_IMAGE_MARKERS):
        return False
    class_text = " ".join(classes) if isinstance(classes, list) else str(classes or "")
    return "avatar" not in class_text.lower()


def extract_images_from_soup(soup: BeautifulSoup, limit: int) -> list[str]:
    """Static fallback matching the original script's topic-detail image rules."""
    images: list[str] = []
    blocks = soup.select("li .topic-detail-image") or soup.select(".topic-detail-image")
    for block in blocks:
        img = block.select_one("img")
        if not img:
            continue
        src = img.get("src") or img.get("data-src") or img.get("data-original") or img.get("data-lazy-src")
        if src and is_valid_topic_image(src, img.get("class")) and src not in images:
            images.append(src)
    return images[:limit]


def extract_info(soup: BeautifulSoup) -> tuple[str, str, str, str, str]:
    price = address = qq = wechat = phone = ""
    for row in soup.find_all("tr"):
        cells = row.find_all(["td", "th"])
        if len(cells) < 2:
            continue
        key = cells[0].get_text(" ", strip=True)
        value = cells[1].get_text(" ", strip=True)
        if "价格" in key:
            price = value
        elif "地址" in key:
            address = value
        elif "QQ" in key.upper():
            qq = value
        elif "微信" in key:
            wechat = value
        elif "电话" in key or "手机" in key:
            phone = value

    text = soup.get_text("\n", strip=True)
    if not qq:
        qq_match = re.search(r"QQ[:：\s]*([0-9]{5,})", text, re.I)
        qq = qq_match.group(1) if qq_match else ""
    if not phone:
        phone_match = re.search(r"(?:电话|手机)[:：\s]*([0-9+\-\s]{7,})", text)
        phone = phone_match.group(1).strip() if phone_match else ""
    return price, address, qq, wechat, phone


def extract_content(soup: BeautifulSoup) -> str:
    content_div = soup.select_one("div.topic-content-detail")
    if content_div:
        return content_div.get_text("\n", strip=True)
    paragraphs = [p.get_text(" ", strip=True) for p in soup.find_all("p")]
    return "\n".join(p for p in paragraphs if p)


def normalize_image_url(page: Any, value: str) -> str:
    if not value:
        return ""
    return page.evaluate("url => new URL(url, location.href).href", value)


def scroll_and_wait_for_lazy_images(page: Any, wait_selector: str, timeout_ms: int) -> None:
    """Trigger lazy loading by scrolling topic image containers into view."""
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError

    page.wait_for_load_state("domcontentloaded", timeout=timeout_ms)
    try:
        page.wait_for_load_state("networkidle", timeout=min(timeout_ms, 8_000))
    except PlaywrightTimeoutError:
        pass

    try:
        page.wait_for_selector(wait_selector, timeout=timeout_ms)
    except PlaywrightTimeoutError:
        pass

    locators = [page.locator(selector) for selector in (".topic-detail-image", ".el-image", "img")]
    for locator in locators:
        count = min(locator.count(), 200)
        if not count:
            continue
        for index in range(count):
            try:
                locator.nth(index).scroll_into_view_if_needed(timeout=2_000)
                page.wait_for_timeout(120)
            except PlaywrightTimeoutError:
                continue
        break

    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    page.wait_for_timeout(500)
    page.evaluate("window.scrollTo(0, 0)")
    page.wait_for_timeout(200)


def extract_image_urls_from_page(page: Any, limit: int) -> list[str]:
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError

    urls: list[str] = []
    for selector in IMAGE_SELECTORS:
        for image in page.locator(selector).all():
            for attr in ("src", "data-src", "data-original", "data-lazy-src"):
                try:
                    value = image.get_attribute(attr)
                    classes = image.get_attribute("class")
                except PlaywrightTimeoutError:
                    value = classes = None
                normalized = normalize_image_url(page, value or "") if value else ""
                if is_valid_topic_image(normalized, classes) and normalized not in urls:
                    urls.append(normalized)
                    if len(urls) >= limit:
                        return urls
    return urls


def parse_rendered_topic(browser: Any, url: str, timeout_ms: int, image_limit: int) -> TopicResult:
    page = browser.new_page(
        viewport={"width": 1365, "height": 900},
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
        ),
        extra_http_headers=HEADERS,
    )
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
        scroll_and_wait_for_lazy_images(page, DEFAULT_WAIT_SELECTOR, timeout_ms)
        html = page.content()
        soup = BeautifulSoup(html, "html.parser")
        title_tag = soup.find("h1") or soup.find("title")
        title = title_tag.get_text(" ", strip=True) if title_tag else "无标题"
        price, address, qq, wechat, phone = extract_info(soup)
        content = extract_content(soup)
        image_urls = extract_image_urls_from_page(page, image_limit)
        if not image_urls:
            image_urls = extract_images_from_soup(soup, image_limit)
        return TopicResult(
            topic_id=parse_topic_id(url),
            url=url,
            ok=True,
            title=clean_text(title),
            price=clean_text(price),
            address=clean_text(address),
            qq=clean_text(qq),
            wechat=clean_text(wechat),
            phone=clean_text(phone),
            content=clean_text(content),
            image_count=len(image_urls),
            image_urls="\n".join(image_urls),
        )
    finally:
        page.close()


def parse_static_topic(url: str, image_limit: int) -> TopicResult | None:
    response = request_url(url)
    if not response:
        return None
    soup = BeautifulSoup(response.text, "html.parser")
    title_tag = soup.find("h1")
    title = title_tag.get_text(strip=True) if title_tag else "无标题"
    price, address, qq, wechat, phone = extract_info(soup)
    content = extract_content(soup)
    images = extract_images_from_soup(soup, image_limit)
    return TopicResult(
        topic_id=parse_topic_id(url),
        url=url,
        ok=True,
        title=clean_text(title),
        price=clean_text(price),
        address=clean_text(address),
        qq=clean_text(qq),
        wechat=clean_text(wechat),
        phone=clean_text(phone),
        content=clean_text(content),
        image_count=len(images),
        image_urls="\n".join(images),
    )


def crawl_topic(browser: Any, url: str, timeout_ms: int, retries: int, image_limit: int, render: bool) -> TopicResult:
    last_error = ""
    for attempt in range(1, retries + 2):
        try:
            result = parse_rendered_topic(browser, url, timeout_ms, image_limit) if render else parse_static_topic(url, image_limit)
            if not result:
                raise RuntimeError("request failed")
            if not result.title and not result.image_urls:
                raise RuntimeError("no title and no image URLs found")
            return result
        except Exception as exc:  # noqa: BLE001 - keep crawler running and report per URL.
            last_error = f"attempt {attempt}: {exc}"
            if attempt <= retries:
                time.sleep(1.5 * attempt + random.random())
    return TopicResult(topic_id=parse_topic_id(url), url=url, ok=False, error=clean_text(last_error))


def discover_topic_links(page_url: str) -> tuple[list[str], str | None]:
    response = request_url(page_url)
    if not response:
        return [], "list request failed"
    soup = BeautifulSoup(response.text, "html.parser")
    links: list[str] = []
    for anchor in soup.select("a[href^='/topic/'], a[href*='/topic/']"):
        href = anchor.get("href", "")
        path = urljoin(BASE_DOMAIN, href)
        if re.search(r"/topic/\d+$", path) and path not in links:
            links.append(path)
    return links, None


def crawl_list_page(page_url: str) -> tuple[str, list[str], str | None]:
    print(f"\n🚀 {page_url}")
    links, error = discover_topic_links(page_url)
    if error:
        print("❌ 列表失败")
    else:
        print(f"📄 帖子数量 {len(links)}")
    return page_url, links, error


def download_image(img_url: str, title: str, index: int, image_dir: Path) -> Path | None:
    try:
        image_dir.mkdir(parents=True, exist_ok=True)
        suffix = Path(img_url.split("?", 1)[0]).suffix or ".jpg"
        filename = clean_filename(f"{title}_{index}{suffix}")
        path = image_dir / filename
        if path.exists():
            return path
        response = session.get(img_url, timeout=15)
        response.raise_for_status()
        path.write_bytes(response.content)
        return path
    except requests.RequestException:
        return None
    except OSError:
        return None


def row_from_result(result: TopicResult) -> list[str]:
    return [
        clean_text(result.title),
        clean_text(result.price),
        clean_text(result.address),
        clean_text(result.qq),
        clean_text(result.wechat),
        clean_text(result.phone),
        clean_text(result.content),
        clean_text(result.url),
        "",
    ]


def save_excel_with_images(results: Sequence[TopicResult], filename: Path, image_dir: Path, embed_images: bool) -> None:
    filename.parent.mkdir(parents=True, exist_ok=True)
    if filename.exists():
        workbook = load_workbook(filename)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["标题", "价格", "地址", "QQ", "微信", "电话", "正文", "链接", "图片"])

    for result in results:
        row_index = worksheet.max_row + 1
        worksheet.append(row_from_result(result))
        if not embed_images:
            worksheet.cell(row=row_index, column=9).value = "\n".join(result.image_list)
            continue
        column = 9
        for index, image_url in enumerate(result.image_list, 1):
            image_path = download_image(image_url, result.title, index, image_dir)
            if not image_path:
                continue
            try:
                image = XLImage(str(image_path))
                image.width = 100
                image.height = 100
                worksheet.add_image(image, worksheet.cell(row=row_index, column=column).coordinate)
                column += 1
            except Exception:  # noqa: BLE001 - bad image bytes should not stop the crawl.
                continue

    workbook.save(filename)
    print(f"💾 写入 {len(results)} 条（{'含图片' if embed_images else '图片URL'}）")


def save_results(results: Sequence[TopicResult], output: Path, image_dir: Path, embed_images: bool) -> None:
    rows = [asdict(result) for result in results]
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.suffix.lower() == ".json":
        output.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    elif output.suffix.lower() == ".csv":
        with output.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=list(TopicResult.__dataclass_fields__.keys()))
            writer.writeheader()
            writer.writerows(rows)
    elif embed_images:
        save_excel_with_images(results, output, image_dir, embed_images=True)
    else:
        pd.DataFrame(rows).to_excel(output, index=False)


def append_results(results: Sequence[TopicResult], output: Path, image_dir: Path, embed_images: bool) -> None:
    if not results:
        return
    if output.suffix.lower() == ".xlsx" and embed_images:
        save_excel_with_images(results, output, image_dir, embed_images=True)
    else:
        save_results(results, output, image_dir, embed_images=False)


def iter_direct_urls(args: argparse.Namespace) -> list[str]:
    urls: list[str] = []
    if args.url:
        urls.extend(args.url)
    if args.url_file:
        urls.extend(line.strip() for line in Path(args.url_file).read_text(encoding="utf-8").splitlines())
    if args.start_id is not None and args.end_id is not None:
        urls.extend(build_topic_url(topic_id) for topic_id in range(args.start_id, args.end_id + 1))
    return [url for url in dict.fromkeys(urls) if url]


def collect_urls_from_pages(args: argparse.Namespace) -> tuple[list[str], set[str]]:
    progress_file = Path(args.progress_file)
    done_pages = load_progress(progress_file) if args.resume else set()
    page_urls = [build_page_url(args.start_url, page_num) for page_num in range(1, args.total_pages + 1)]
    pending_pages = [url for url in page_urls if url not in done_pages]
    print(f"🧵 待爬 {len(pending_pages)} 页")

    all_links: list[str] = []
    for index in range(0, len(pending_pages), args.page_threads):
        batch = pending_pages[index:index + args.page_threads]
        print(f"\n🔥 列表批次 {index // args.page_threads + 1}")
        with ThreadPoolExecutor(max_workers=args.page_threads) as executor:
            futures = {executor.submit(crawl_list_page, url): url for url in batch}
            for future in as_completed(futures):
                page_url, links, error = future.result()
                if not error:
                    done_pages.add(page_url)
                    save_progress(done_pages, progress_file)
                all_links.extend(links)
        time.sleep(args.page_delay)
    return [url for url in dict.fromkeys(all_links)], done_pages


def crawl_urls(args: argparse.Namespace, urls: Sequence[str]) -> list[TopicResult]:
    if not urls:
        return []

    from playwright.sync_api import sync_playwright

    results: list[TopicResult] = []
    failed: list[str] = []
    output = Path(args.output)
    image_dir = Path(args.image_dir)

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=not args.headful)
        try:
            for index, url in enumerate(urls, 1):
                print(f"[{index}/{len(urls)}] crawling {url}")
                result = crawl_topic(browser, url, args.timeout, args.retries, args.image_limit, not args.static_only)
                if result.ok:
                    results.append(result)
                    if output.suffix.lower() == ".xlsx" and args.embed_images:
                        append_results([result], output, image_dir, args.embed_images)
                    else:
                        save_results(results, output, image_dir, embed_images=False)
                    print(f"  OK: images={result.image_count} title={result.title}")
                else:
                    failed.append(url)
                    print(f"  FAIL: {result.error}")
                time.sleep(args.topic_delay)

            if failed:
                print(f"\n🔁 重试失败帖子 {len(failed)} 条")
                retry_failed: list[str] = []
                for url in failed:
                    result = crawl_topic(browser, url, args.timeout, args.retries, args.image_limit, not args.static_only)
                    if result.ok:
                        results.append(result)
                        if output.suffix.lower() == ".xlsx" and args.embed_images:
                            append_results([result], output, image_dir, args.embed_images)
                        else:
                            save_results(results, output, image_dir, embed_images=False)
                    else:
                        retry_failed.append(url)
                failed = retry_failed
        finally:
            browser.close()

    if failed:
        Path(args.failed_file).write_text("\n".join(failed) + "\n", encoding="utf-8")
    return results


def parse_args(argv: Iterable[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="XC8866 rendered lazy-image crawler with list-page mode")
    input_group = parser.add_argument_group("topic inputs")
    input_group.add_argument("--url", action="append", help="Topic URL. Can be supplied multiple times.")
    input_group.add_argument("--url-file", help="Text file containing one topic URL per line.")
    input_group.add_argument("--start-id", type=int, help="Start topic ID, inclusive.")
    input_group.add_argument("--end-id", type=int, help="End topic ID, inclusive.")

    list_group = parser.add_argument_group("list-page inputs")
    list_group.add_argument("--start-url", help="List/category URL used with --total-pages.")
    list_group.add_argument("--total-pages", type=int, help="Total list pages to crawl.")
    list_group.add_argument("--page-threads", type=int, default=4, help="Concurrent list-page fetchers.")
    list_group.add_argument("--resume", action="store_true", help="Skip list pages saved in progress file.")
    list_group.add_argument("--progress-file", default=PROGRESS_FILE, help="Progress checkpoint JSON path.")

    parser.add_argument("--threads", type=int, default=6, help="Compatibility option; topic rendering stays sequential for browser stability.")
    parser.add_argument("--output", default="result.xlsx", help="Output .xlsx, .csv, or .json file.")
    parser.add_argument("--image-dir", default=IMAGE_DIR, help="Downloaded image directory for embedded Excel mode.")
    parser.add_argument("--failed-file", default="failed_links.txt", help="Failed topic URL output path.")
    parser.add_argument("--embed-images", action="store_true", help="Download topic images and embed them into .xlsx output.")
    parser.add_argument("--static-only", action="store_true", help="Do not render with Playwright; use static HTML fallback only.")
    parser.add_argument("--headful", action="store_true", help="Show the browser window for debugging.")
    parser.add_argument("--timeout", type=int, default=30_000, help="Per-page timeout in milliseconds.")
    parser.add_argument("--retries", type=int, default=2, help="Retry count after the first attempt.")
    parser.add_argument("--image-limit", type=int, default=4, help="Maximum topic images to keep, matching the original script default.")
    parser.add_argument("--page-delay", type=float, default=1.0, help="Delay between list-page batches.")
    parser.add_argument("--topic-delay", type=float, default=0.2, help="Delay between topic renders.")
    return parser.parse_args(argv)


def main(argv: Iterable[str] | None = None) -> int:
    args = parse_args(argv)
    urls = iter_direct_urls(args)
    if args.start_url or args.total_pages:
        if not args.start_url or not args.total_pages:
            raise SystemExit("--start-url and --total-pages must be provided together.")
        page_urls, _done_pages = collect_urls_from_pages(args)
        urls.extend(page_urls)
    urls = [url for url in dict.fromkeys(urls) if url]
    if not urls:
        raise SystemExit("Please provide --url, --url-file, --start-id/--end-id, or --start-url/--total-pages.")

    results = crawl_urls(args, urls)
    print(f"Saved {len(results)} successful rows to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
