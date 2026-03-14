import argparse
import json
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

BASE_DOMAIN = "https://xc8866.com"
PROGRESS_FILE = "progress.json"

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Referer": BASE_DOMAIN,
}

session = requests.Session()
session.headers.update(HEADERS)

ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


# ======================
# 文本清理
# ======================
def clean_text(text):
    if not text:
        return ""
    return ILLEGAL_CHARACTERS_RE.sub("", str(text))


# ======================
# 固定分页替换
# ======================
def build_page_url(base_url, page_num):
    if "page=" in base_url:
        return re.sub(r"page=\d+", f"page={page_num}", base_url)

    if "?" in base_url:
        return base_url + f"&page={page_num}"
    return base_url + f"?page={page_num}"


def extract_page_num(url):
    match = re.search(r"[?&]page=(\d+)", url)
    return int(match.group(1)) if match else 1


# ======================
# 进度记录
# ======================
def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def save_progress(done_pages):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(list(done_pages), f)


# ======================
# Excel 写入（批量）
# ======================
def save_excel(data_list, filename="result.xlsx"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["标题", "价格", "地址", "QQ", "微信", "电话", "正文", "链接"])

    for data in data_list:
        ws.append(
            [
                clean_text(data["标题"]),
                clean_text(data["价格"]),
                clean_text(data["地址"]),
                clean_text(data["QQ"]),
                clean_text(data["微信"]),
                clean_text(data["电话"]),
                clean_text(data["正文"]),
                clean_text(data["链接"]),
            ]
        )

    wb.save(filename)
    print(f"💾 批量写入 {len(data_list)} 条")


# ======================
# 请求
# ======================
def request(url):
    try:
        resp = session.get(url, timeout=15)
        resp.raise_for_status()
        if not resp.encoding:
            resp.encoding = resp.apparent_encoding
        return resp
    except requests.RequestException:
        return None


def detect_total_pages(start_url):
    resp = request(start_url)
    if not resp:
        return extract_page_num(start_url)

    soup = BeautifulSoup(resp.text, "html.parser")
    page_numbers = [extract_page_num(start_url)]

    for a in soup.select("a[href*='page=']"):
        href = a.get("href", "")
        if href.startswith("/"):
            href = BASE_DOMAIN + href
        page_numbers.append(extract_page_num(href))

    return max(page_numbers) if page_numbers else extract_page_num(start_url)


# ======================
# 解析帖子
# ======================
def parse_thread(url):
    resp = request(url)
    if not resp:
        return None

    soup = BeautifulSoup(resp.text, "html.parser")

    title_tag = soup.find("h1")
    title = title_tag.get_text(strip=True) if title_tag else "无标题"

    price = qq = wechat = phone = address = ""

    for tr in soup.select("tr"):
        tds = tr.find_all("td")
        if len(tds) >= 2:
            key = tds[0].get_text(strip=True)
            value = tds[1].get_text(strip=True)

            if "价格" in key:
                price = value
            elif "地址" in key:
                address = value
            elif "QQ" in key:
                qq = value
            elif "微信" in key:
                wechat = value
            elif "电话" in key or "手机" in key:
                phone = value

    content_div = soup.select_one("div.topic-content-detail")
    content = content_div.get_text("\n", strip=True) if content_div else ""

    return {
        "标题": title,
        "价格": price,
        "地址": address,
        "QQ": qq,
        "微信": wechat,
        "电话": phone,
        "正文": content,
        "链接": url,
    }


# ======================
# 抓列表页
# ======================
def crawl_page(page_url, thread_workers):
    print(f"\n🚀 {page_url}")

    resp = request(page_url)
    if not resp:
        print("❌ 列表失败")
        return [], []

    soup = BeautifulSoup(resp.text, "html.parser")

    links = []
    for a in soup.select("a[href^='/topic/']"):
        href = a["href"]
        if re.match(r"^/topic/\d+$", href):
            links.append(BASE_DOMAIN + href)

    links = list(set(links))
    print(f"📄 帖子数量 {len(links)}")

    results = []
    failed = []

    start_time = time.time()

    with ThreadPoolExecutor(max_workers=thread_workers) as executor:
        futures = {executor.submit(parse_thread, link): link for link in links}

        for future in as_completed(futures):
            data = future.result()

            if data:
                results.append(data)
            else:
                failed.append(futures[future])

    elapsed = time.time() - start_time
    speed = len(results) / elapsed if elapsed else 0

    print(f"⚡ 速度 {speed:.2f} 帖子/秒")

    return results, failed


# ======================
# 主流程
# ======================
def crawl_pages(start_url, total_pages, page_threads, thread_workers):
    start_page = extract_page_num(start_url)
    done_pages = load_progress()
    all_pages = []

    for i in range(start_page, total_pages + 1):
        url = build_page_url(start_url, i)
        if url not in done_pages:
            all_pages.append(url)

    print(f"🧵 待爬 {len(all_pages)} 页")

    retry_queue = []

    for i in range(0, len(all_pages), page_threads):
        batch = all_pages[i : i + page_threads]
        batch_data = []

        print(f"\n🔥 批次 {i // page_threads + 1}")

        with ThreadPoolExecutor(max_workers=page_threads) as executor:
            futures = {executor.submit(crawl_page, url, thread_workers): url for url in batch}

            for future in as_completed(futures):
                page_data, failed = future.result()
                batch_data.extend(page_data)
                retry_queue.extend(failed)
                done_pages.add(futures[future])

        # 自动重试
        if retry_queue:
            print(f"🔄 重试 {len(retry_queue)} 个失败帖子")
            retry_results = []

            with ThreadPoolExecutor(max_workers=thread_workers) as executor:
                futures = [executor.submit(parse_thread, url) for url in retry_queue]

                for future in as_completed(futures):
                    data = future.result()
                    if data:
                        retry_results.append(data)

            batch_data.extend(retry_results)
            retry_queue.clear()

        if batch_data:
            save_excel(batch_data)

        save_progress(done_pages)

        print("✅ 批次完成")
        time.sleep(1)


# ======================
# main
# ======================
def main():
    parser = argparse.ArgumentParser(description="从 xc8866 列表页开始爬取帖子信息")
    parser.add_argument("--start-url", default="https://xc8866.com/?page=1")
    parser.add_argument("--total-pages", type=int)
    parser.add_argument("--page-threads", type=int, default=4)
    parser.add_argument("--threads", type=int, default=6)

    args = parser.parse_args()

    total_pages = args.total_pages or detect_total_pages(args.start_url)
    print(f"📚 总页数 {total_pages}（起始 URL: {args.start_url}）")

    crawl_pages(args.start_url, total_pages, args.page_threads, args.threads)


if __name__ == "__main__":
    main()
